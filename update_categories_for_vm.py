#!/usr/bin/env python3
"""
SYNOPSIS:
    Script Name: update_categories_for_vm.py
    Author: hardev@nutanix.com + Co-Pilot
    Date: October 2025
    Version: 1.0
    Purpose:
    A script to update VM categories using Nutanix API v4 REST calls.
    Reads configuration from vars.txt and processes VMs from an Excel file.

NB:
    This script is provided "AS IS" without warranty of any kind.
    Use of this script is at your own risk.
    The author(s) make no representations or warranties, express or implied,
    regarding the script’s functionality, fitness for a particular purpose,
    or reliability.

    By using this script, you agree that you are solely responsible
    for any outcomes, including loss of data, system issues, or
    other damages that may result from its execution.
    No support or maintenance is provided.

NOTES:
    You may copy, edit, customize and use as needed.
    Test thoroughly in a safe environment before deploying to production systems.
"""

import os
import sys
import json
import uuid
import requests
import pandas as pd
from datetime import datetime
from base64 import b64encode
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import urllib3

# Disable SSL warnings for self-signed certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def read_vars_file():
    """Read configuration variables from vars.txt file"""
    vars_file = os.path.join(os.getcwd(), 'files', 'vars.txt')
    
    if not os.path.exists(vars_file):
        print(f"Error: vars.txt file not found at {vars_file}")
        sys.exit(1)
    
    vars_dict = {}
    with open(vars_file, 'r') as f:
        for line in f:
            line = line.strip()
            if '=' in line:
                key, value = line.split('=', 1)
                vars_dict[key] = value
    
    return vars_dict

def create_auth_header(username, password):
    """Create Basic Authentication header"""
    credentials = f"{username}:{password}"
    encoded_credentials = b64encode(credentials.encode('ascii')).decode('ascii')
    return {'Authorization': f'Basic {encoded_credentials}'}

def make_get_request(base_url, uri_get, headers):
    """Make GET request to retrieve VM information and ETag"""
    full_url = f"{base_url}/{uri_get}"
    
    try:
        print(f"Making GET request to: {full_url}")
        response = requests.get(
            full_url,
            headers=headers,
            verify=False,  # Skip certificate verification
            timeout=30
        )
        
        # Pretty print the returned JSON
        print("GET Response JSON:")
        print(json.dumps(response.json(), indent=2))
        
        # Extract ETag from response headers
        vm_etag = response.headers.get('ETag', '')
        print(f"VM ETag: {vm_etag}")
        
        return vm_etag, response.json()
        
    except requests.exceptions.RequestException as e:
        print(f"Error making GET request: {e}")
        return None, None

def build_categories_payload(category_uuids):
    """Build JSON payload for categories association"""
    categories = []
    
    for uuid_str in category_uuids:
        uuid_str = uuid_str.strip()
        if uuid_str:  # Only add non-empty UUIDs
            categories.append({"extId": uuid_str})
    
    payload = {"categories": categories}
    return payload

def make_post_request(base_url, uri_post, headers, payload):
    """Make POST request to associate categories with VM"""
    full_url = f"{base_url}/{uri_post}"
    
    try:
        print(f"Making POST request to: {full_url}")
        print("POST Payload:")
        print(json.dumps(payload, indent=2))
        
        response = requests.post(
            full_url,
            headers=headers,
            json=payload,
            verify=False,  # Skip certificate verification
            timeout=30
        )
        
        # Pretty print the returned response
        print("POST Response:")
        print(f"Status Code: {response.status_code}")
        print(f"Headers: {dict(response.headers)}")
        
        if response.text:
            try:
                print("Response JSON:")
                print(json.dumps(response.json(), indent=2))
            except json.JSONDecodeError:
                print("Response Text:")
                print(response.text)
        
        return response
        
    except requests.exceptions.RequestException as e:
        print(f"Error making POST request: {e}")
        return None

def update_excel_status(file_path, sheet_name, row_index, status, timestamp):
    """Update the Excel file with status and timestamp"""
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
        
        # Find the column indices for STATUS and TIMESTAMP
        status_col = None
        timestamp_col = None
        
        for col in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            if header and "STATUS OF UPDATE" in str(header).upper():
                status_col = col
            elif header and "TIMESTAMP" in str(header).upper():
                timestamp_col = col
        
        if status_col:
            # Update status with green background and white bold text
            cell = worksheet.cell(row=row_index + 2, column=status_col)  # +2 because Excel is 1-indexed and we skip header
            cell.value = status
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        if timestamp_col:
            # Update timestamp
            cell = worksheet.cell(row=row_index + 2, column=timestamp_col)
            cell.value = timestamp
        
        workbook.save(file_path)
        print(f"Updated Excel file with status: {status} and timestamp: {timestamp}")
        
    except Exception as e:
        print(f"Error updating Excel file: {e}")

def main():
    """Main function to process VMs and update categories"""
    print("Starting update_categories_for_vm.py script")
    
    # Read configuration variables
    vars_dict = read_vars_file()
    base_url = vars_dict.get('baseUrl', '')
    username = vars_dict.get('username', '')
    password = vars_dict.get('password', '')
    
    if not all([base_url, username, password]):
        print("Error: Missing required variables in vars.txt (baseUrl, username, password)")
        sys.exit(1)
    
    print(f"Using base URL: {base_url}")
    print(f"Using username: {username}")
    
    # Create authentication headers
    auth_headers = create_auth_header(username, password)
    
    # Read Excel file
    excel_file = os.path.join(os.getcwd(), 'scratch', 'VMsToUpdate-PROD.xlsx')
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found at {excel_file}")
        sys.exit(1)
    
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file, sheet_name='ToUpdate')
        print(f"Loaded Excel file with {len(df)} rows")
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)
    
    # Process each row
    for index, row in df.iterrows():
        # Check if the match status is "OK"
        match_status = str(row.get('VM Name/extId & Category exId(s) Match', '')).strip()
        
        if match_status.upper() == 'OK':
            vm_name = str(row.get('VM Name', '')).strip()
            vm_extid = str(row.get('VM extId', '')).strip()
            category_uuids_str = str(row.get('Category UUID(s)', '')).strip()
            
            if not all([vm_name, vm_extid, category_uuids_str]):
                print(f"Skipping row {index + 1}: Missing required data")
                continue
            
            print(f"\n{'='*60}")
            print(f"Processing VM: {vm_name} (extId: {vm_extid})")
            print(f"Category UUIDs: {category_uuids_str}")
            
            # Parse category UUIDs (comma-separated)
            category_uuids = [uuid.strip() for uuid in category_uuids_str.split(',') if uuid.strip()]
            
            if not category_uuids:
                print(f"No valid category UUIDs found for VM {vm_name}")
                continue
            
            # Initialize URI variables
            uri_get = f"vmm/v4.1/ahv/config/vms/{vm_extid}"
            uri_post = f"vmm/v4.1/ahv/config/vms/{vm_extid}/$actions/associate-categories"
            
            # Generate REST UUID
            rest_uuid = str(uuid.uuid4())
            
            # 1st REST call: GET
            print(f"\n1. Making GET request...")
            vm_etag, get_response = make_get_request(base_url, uri_get, auth_headers)
            
            if not vm_etag:
                print(f"Failed to get ETag for VM {vm_name}, skipping...")
                continue
            
            # Build JSON payload for POST
            payload = build_categories_payload(category_uuids)
            
            # Prepare POST headers
            post_headers = auth_headers.copy()
            post_headers.update({
                'If-Match': vm_etag,
                'NTNX-Request-Id': rest_uuid,
                'Content-Type': 'application/json'
            })
            
            # 2nd REST call: POST
            print(f"\n2. Making POST request...")
            post_response = make_post_request(base_url, uri_post, post_headers, payload)
            
            if post_response and post_response.status_code == 202:
                # Success - extract ETag from response headers
                vm_etag_updated = post_response.headers.get('ETag', '')
                print(f"ACCEPTED: Updated VM ETag: {vm_etag_updated}")
                
                # Update Excel with success status
                timestamp = datetime.now().strftime("%d%m%Y-%H%M")
                update_excel_status(excel_file, 'ToUpdate', index, 'ACCEPTED', timestamp)
                
            else:
                status_code = post_response.status_code if post_response else 'N/A'
                print(f"FAILED: POST request failed with status code: {status_code}")
                
                # Update Excel with failure status
                timestamp = datetime.now().strftime("%d%m%Y-%H%M")
                update_excel_status(excel_file, 'ToUpdate', index, f'FAILED ({status_code})', timestamp)
        
        else:
            print(f"Skipping row {index + 1}: Match status is '{match_status}', not 'OK'")
    
    print(f"\n{'='*60}")
    print("Script completed!")

if __name__ == "__main__":
    main()